"""
Matrix spreadsheet importer.

Opens an Excel file from the HQA network drive, lets the user pick a
worksheet and a data row, then maps the column headers to FR field keys
and returns the pre-filled values to NewReportDialog.

Requires openpyxl:
    pip install openpyxl
"""

import os

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QComboBox,
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QVBoxLayout,
    QFormLayout,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HQA_BASE = r"\\uslafvs001038\HQA\TEST DATA\00-GE LAB IN PROCESS"

# ---------------------------------------------------------------------------
# Column-header → FR db_key mapping
# Keys are normalized (lowercase, stripped, collapsed spaces).
# A single FR field may match many Excel header variations.
# ---------------------------------------------------------------------------

_HEADER_MAP: dict[str, str] = {
    # ── Project / test info ──────────────────────────────────────────────
    "project":                  "Project",
    "project name":             "Project",
    "project number":           "Project_Number",
    "proj number":              "Project_Number",
    "proj #":                   "Project_Number",
    "project #":                "Project_Number",
    "project num":              "Project_Number",
    "fw version":               "FW Ver",
    "fw ver":                   "FW Ver",
    "firmware":                 "FW Ver",
    "firmware version":         "FW Ver",
    "fwver":                    "FW Ver",
    "test matrix id":           "Test_Matrix_ID",
    "matrix id":                "Test_Matrix_ID",
    "test":                     "Test",
    "test name":                "Test",
    "test type":                "Test_Type",
    "level":                    "Level",
    "date failed":              "Date Failed",
    "fail date":                "Date Failed",
    "failure date":             "Date Failed",
    "date":                     "Date Failed",
    "tested by":                "Tested By",
    "tester":                   "Tested By",
    "assigned to":              "Assigned To",
    "project lead":             "Assigned To",
    "lead engineer":            "Assigned To",
    "engineer":                 "Assigned To",
    "eut type":                 "EUT_TYPE",
    "equipment type":           "EUT_TYPE",
    "eut":                      "EUT_TYPE",
    # ── Meter info ───────────────────────────────────────────────────────
    "meter":                    "Meter",
    "meter model":              "Meter",
    "meter type":               "Meter_Type",
    "meter manufacturer":       "Meter_Manufacturer",
    "manufacturer":             "Meter_Manufacturer",
    "mfr":                      "Meter_Manufacturer",
    "meter mfr":                "Meter_Manufacturer",
    "sub type":                 "Meter_SubType",
    "subtype":                  "Meter_SubType",
    "meter sub type":           "Meter_SubType",
    "meter subtype":            "Meter_SubType",
    "sub type ii":              "Meter_SubTypeII",
    "subtype ii":               "Meter_SubTypeII",
    "meter sub type ii":        "Meter_SubTypeII",
    "serial number":            "Meter_Serial_Number",
    "serial #":                 "Meter_Serial_Number",
    "serial no":                "Meter_Serial_Number",
    "sn":                       "Meter_Serial_Number",
    "meter serial":             "Meter_Serial_Number",
    "meter sn":                 "Meter_Serial_Number",
    "meter serial number":      "Meter_Serial_Number",
    "form":                     "Form",
    "form factor":              "Form",
    "meter form":               "Form",
    "base":                     "Meter_Base",
    "meter base":               "Meter_Base",
    "voltage":                  "Meter_Voltage",
    "meter voltage":            "Meter_Voltage",
    "dsp rev":                  "Meter_DSP_Rev",
    "meter dsp rev":            "Meter_DSP_Rev",
    "pcba":                     "Meter_PCBA",
    "pcba number":              "Meter_PCBA",
    "meter pcba":               "Meter_PCBA",
    "pcba rev":                 "Meter_PCBA_Rev",
    "meter pcba rev":           "Meter_PCBA_Rev",
    "software":                 "Meter_Software",
    "meter software":           "Meter_Software",
    "software rev":             "Meter_Software_Rev",
    "meter software rev":       "Meter_Software_Rev",
    "sw rev":                   "Meter_Software_Rev",
    "meter notes":              "Meter_Notes",
    # ── AMR / Radio info ─────────────────────────────────────────────────
    "amr":                      "AMR",
    "amr model":                "AMR",
    "radio":                    "AMR",
    "amr rev":                  "AMR Rev",
    "amr revision":             "AMR Rev",
    "amr manufacturer":         "AMR_Manufacturer",
    "radio manufacturer":       "AMR_Manufacturer",
    "amr type":                 "AMR_Type",
    "radio type":               "AMR_Type",
    "amr sub type":             "AMR_SUBType",
    "amr subtype":              "AMR_SUBType",
    "radio subtype":            "AMR_SUBType",
    "amr serial":               "AMR_SN",
    "amr sn":                   "AMR_SN",
    "amr serial number":        "AMR_SN",
    "radio sn":                 "AMR_SN",
    "amr notes":                "AMR_Notes",
    "radio notes":              "AMR_Notes",
    "amr software":             "AMR_Software",
    "amr sw":                   "AMR_Software",
    "amr software rev":         "AMR_Software_Rev",
    "amr sw rev":               "AMR_Software_Rev",
    "amr pcba":                 "AMR_PCBA",
    "amr pcba rev":             "AMR_PCBA_Rev",
    "amr voltage":              "AMR_Voltage",
    "amr ip":                   "AMR_IP_LAN_ID",
    "amr ip lan id":            "AMR_IP_LAN_ID",
    "ip lan id":                "AMR_IP_LAN_ID",
    # ── Failure / notes ──────────────────────────────────────────────────
    "failure description":      "Failure Description",
    "failure":                  "Failure Description",
    "description":              "Failure Description",
    "corrective action":        "Corrective Action",
    "engineering notes":        "Engineering Notes",
    "tcc comments":             "TCC Comments",
    "test equipment id":        "Test_Equipment_ID",
    "equipment id":             "Test_Equipment_ID",
}

# Ordered list of Excel header keys used to build the row preview in the dialog
_PREVIEW_HINTS = [
    "project", "project name", "meter", "meter model",
    "meter type", "serial number", "sn", "amr", "test",
]


# ---------------------------------------------------------------------------
# Excel reading helpers
# ---------------------------------------------------------------------------

def _normalize(s: str) -> str:
    """Lowercase, strip, collapse whitespace."""
    return " ".join(str(s).strip().lower().split())


def read_workbook(path: str) -> dict:
    """
    Read an Excel file and return:
      {
        "sheets": [str, ...],
        "data":   {sheet_name: [row_dict, ...]}
      }

    Each row_dict maps the column header string to the cell value string.
    Empty rows are skipped.

    Raises:
        ImportError  — openpyxl not installed
        ValueError   — file could not be opened
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required to read Excel files.\n\n"
            "Install it by running:\n    pip install openpyxl"
        )
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not open workbook:\n{exc}")

    result: dict = {"sheets": wb.sheetnames, "data": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result["data"][name] = []
            continue
        headers = [
            str(c).strip() if c is not None else f"_col{i}"
            for i, c in enumerate(rows[0])
        ]
        sheet_rows = []
        for row_vals in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row_vals):
                continue
            row_dict = {
                headers[i]: str(v).strip()
                for i, v in enumerate(row_vals)
                if i < len(headers) and headers[i] and not headers[i].startswith("_col")
                and v is not None and str(v).strip()
            }
            if row_dict:
                sheet_rows.append(row_dict)
        result["data"][name] = sheet_rows

    wb.close()
    return result


def map_row_to_fr(row: dict) -> dict:
    """
    Map a row dict (Excel header → cell value) to FR db_keys.
    Only keys with non-empty string values are included.
    """
    out: dict = {}
    for header, value in row.items():
        fr_key = _HEADER_MAP.get(_normalize(header))
        if fr_key and value:
            out[fr_key] = value
    return out


def _row_preview(row: dict, row_num: int) -> str:
    """Build a short human-readable summary line for a data row."""
    parts = []
    norm_row = {_normalize(k): v for k, v in row.items()}
    for hint in _PREVIEW_HINTS:
        v = norm_row.get(hint, "")
        if v and v not in parts:
            parts.append(v)
        if len(parts) >= 3:
            break
    label = f"Row {row_num}"
    if parts:
        label += f"  —  {' | '.join(parts)}"
    return label


# ---------------------------------------------------------------------------
# Import dialog
# ---------------------------------------------------------------------------

class MatrixImportDialog(QDialog):
    """
    Two-step import dialog:
      1. Browse to an Excel file on the HQA drive
      2. Select the worksheet and which data row to import

    After accept(), call get_mapped_fields() to retrieve the dict of
    FR db_key → str_value ready to pre-fill NewReportDialog.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._workbook_data: dict | None = None
        self._mapped_fields: dict = {}
        self.setWindowTitle("Import from Matrix Spreadsheet")
        self.resize(580, 440)
        self._build_ui()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setSpacing(10)
        root.setContentsMargins(12, 12, 12, 12)

        # ── File row ────────────────────────────────────────────────────
        file_row = QHBoxLayout()
        self._path_label = QLabel("No file selected")
        self._path_label.setWordWrap(True)
        self._path_label.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred
        )
        self._path_label.setStyleSheet("color: #555; font-style: italic;")
        file_row.addWidget(self._path_label)

        browse_btn = QPushButton("Browse…")
        browse_btn.setObjectName("outline_btn")
        browse_btn.setFixedWidth(80)
        browse_btn.clicked.connect(self._on_browse)
        file_row.addWidget(browse_btn)
        root.addLayout(file_row)

        # ── Sheet selector ───────────────────────────────────────────────
        form = QFormLayout()
        form.setHorizontalSpacing(12)
        self._sheet_combo = QComboBox()
        self._sheet_combo.currentIndexChanged.connect(self._on_sheet_changed)
        form.addRow("Worksheet:", self._sheet_combo)
        root.addLayout(form)

        # ── Row list ─────────────────────────────────────────────────────
        root.addWidget(QLabel("Select the row to import:"))
        self._row_list = QListWidget()
        self._row_list.setMinimumHeight(160)
        self._row_list.itemSelectionChanged.connect(self._on_row_selection_changed)
        root.addWidget(self._row_list, stretch=1)

        # ── Status label ─────────────────────────────────────────────────
        self._status_label = QLabel("")
        self._status_label.setStyleSheet("color: #888; font-size: 8pt;")
        root.addWidget(self._status_label)

        # ── Buttons ──────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self._import_btn = QPushButton("Import Selected Row")
        self._import_btn.setObjectName("save_btn")
        self._import_btn.setEnabled(False)
        self._import_btn.clicked.connect(self._on_import)
        btn_row.addWidget(self._import_btn)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("cancel_btn")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(cancel_btn)
        root.addLayout(btn_row)

    # ------------------------------------------------------------------
    # Slots
    # ------------------------------------------------------------------

    def _on_browse(self):
        start = HQA_BASE if os.path.exists(HQA_BASE) else ""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Matrix Spreadsheet",
            start,
            "Excel Files (*.xlsx *.xlsm *.xls);;All Files (*)",
        )
        if not path:
            return

        try:
            wb = read_workbook(path)
        except ImportError as exc:
            QMessageBox.critical(self, "Missing Library", str(exc))
            return
        except ValueError as exc:
            QMessageBox.critical(self, "File Error", str(exc))
            return

        self._workbook_data = wb
        self._path_label.setText(os.path.basename(path))
        self._path_label.setStyleSheet("color: #222;")

        self._sheet_combo.blockSignals(True)
        self._sheet_combo.clear()
        self._sheet_combo.addItems(wb["sheets"])
        self._sheet_combo.blockSignals(False)
        self._on_sheet_changed()

    def _on_sheet_changed(self):
        self._row_list.clear()
        self._import_btn.setEnabled(False)
        self._status_label.setText("")
        if not self._workbook_data:
            return

        sheet = self._sheet_combo.currentText()
        rows = self._workbook_data["data"].get(sheet, [])

        if not rows:
            self._status_label.setText("No data rows found in this worksheet.")
            return

        for i, row in enumerate(rows, start=1):
            item_text = _row_preview(row, i)
            self._row_list.addItem(item_text)
            self._row_list.item(self._row_list.count() - 1).setData(
                Qt.ItemDataRole.UserRole, row
            )

        self._row_list.setCurrentRow(0)
        self._status_label.setText(f"{len(rows)} data row(s) found.")

    def _on_row_selection_changed(self):
        self._import_btn.setEnabled(self._row_list.currentItem() is not None)

    def _on_import(self):
        item = self._row_list.currentItem()
        if item is None:
            return
        row = item.data(Qt.ItemDataRole.UserRole)
        mapped = map_row_to_fr(row)
        if not mapped:
            QMessageBox.warning(
                self,
                "No Matching Fields",
                "No column headers in the selected row matched known FR fields.\n\n"
                "The spreadsheet column headers may need to match FR field names\n"
                "(e.g. 'Meter Type', 'Serial Number', 'Project Name', etc.).",
            )
            return
        self._mapped_fields = mapped
        self.accept()

    # ------------------------------------------------------------------
    # Result
    # ------------------------------------------------------------------

    def get_mapped_fields(self) -> dict:
        """Return the mapped FR fields after a successful import."""
        return dict(self._mapped_fields)
